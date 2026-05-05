from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import sqlite3, uuid, json, os, time, io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

app = Flask(__name__, static_folder='public')
CORS(app)

DB_PATH = 'gara.db'

def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.executescript('''
        CREATE TABLE IF NOT EXISTS admins (
            id TEXT PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            created_at REAL DEFAULT (strftime('%s','now'))
        );
        CREATE TABLE IF NOT EXISTS tests (
            id TEXT PRIMARY KEY,
            test_code TEXT UNIQUE NOT NULL,
            subject TEXT NOT NULL,
            grade_levels TEXT NOT NULL,
            questions TEXT NOT NULL,
            duration_seconds INTEGER DEFAULT 3600,
            created_by TEXT,
            created_at REAL DEFAULT (strftime('%s','now')),
            active INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS student_sessions (
            id TEXT PRIMARY KEY,
            test_id TEXT NOT NULL,
            student_name TEXT NOT NULL,
            student_surname TEXT NOT NULL,
            grade TEXT NOT NULL,
            school TEXT NOT NULL,
            started_at REAL,
            submitted_at REAL,
            answers TEXT DEFAULT '{}',
            score INTEGER DEFAULT 0,
            max_score INTEGER DEFAULT 100,
            duration_taken REAL,
            wrong_questions TEXT DEFAULT '[]',
            correct_questions TEXT DEFAULT '[]',
            status TEXT DEFAULT 'not_started',
            FOREIGN KEY(test_id) REFERENCES tests(id)
        );
        CREATE TABLE IF NOT EXISTS schools (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL
        );
    ''')
    # Insert default admin
    try:
        c.execute("INSERT OR IGNORE INTO admins (id, username, password) VALUES (?,?,?)",
                  (str(uuid.uuid4()), 'admin', 'admin123'))
    except: pass
    # Insert schools from document
    schools = [
        ('A. SH.', 'Shkolla A. SH.'),
        ('B.', 'Shkolla B.'),
        ('C.', 'Shkolla C.'),
        ('E.', 'Shkolla E.'),
        ('D.', 'Shkolla D.'),
        ('I. A.', 'Shkolla I. A.'),
        ('H. P.', 'Shkolla H. P.'),
        ('L. P.', 'Shkolla L. P.'),
        ('S', 'Shkolla S'),
    ]
    for sid, sname in schools:
        c.execute("INSERT OR IGNORE INTO schools (id, name) VALUES (?,?)", (sid, sname))
    conn.commit()
    conn.close()

# ── AUTH ─────────────────────────────────────────────────────────────────────

@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    data = request.json
    conn = get_db()
    admin = conn.execute("SELECT * FROM admins WHERE username=? AND password=?",
                         (data['username'], data['password'])).fetchone()
    conn.close()
    if admin:
        return jsonify({'success': True, 'adminId': admin['id'], 'username': admin['username']})
    return jsonify({'success': False, 'error': 'Kredencialet janë të gabuara'}), 401

# ── SCHOOLS ──────────────────────────────────────────────────────────────────

@app.route('/api/schools', methods=['GET'])
def get_schools():
    conn = get_db()
    schools = conn.execute("SELECT * FROM schools").fetchall()
    conn.close()
    return jsonify([dict(s) for s in schools])

# ── TESTS ─────────────────────────────────────────────────────────────────────

@app.route('/api/tests', methods=['POST'])
def create_test():
    data = request.json
    tid = str(uuid.uuid4())
    # Generate unique 6-char test code
    test_code = str(uuid.uuid4())[:6].upper()
    conn = get_db()
    # Ensure unique code
    while conn.execute("SELECT id FROM tests WHERE test_code=?", (test_code,)).fetchone():
        test_code = str(uuid.uuid4())[:6].upper()
    conn.execute('''INSERT INTO tests (id, test_code, subject, grade_levels, questions, duration_seconds, created_by)
                    VALUES (?,?,?,?,?,?,?)''',
                 (tid, test_code, data['subject'], json.dumps(data['grade_levels']),
                  json.dumps(data['questions']), data.get('duration_seconds', 3600), data.get('admin_id')))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'test_id': tid, 'test_code': test_code})

@app.route('/api/tests', methods=['GET'])
def get_tests():
    conn = get_db()
    tests = conn.execute("SELECT id, test_code, subject, grade_levels, created_at, active FROM tests ORDER BY created_at DESC").fetchall()
    conn.close()
    result = []
    for t in tests:
        d = dict(t)
        d['grade_levels'] = json.loads(d['grade_levels'])
        result.append(d)
    return jsonify(result)

@app.route('/api/tests/<test_id>', methods=['GET'])
def get_test(test_id):
    conn = get_db()
    t = conn.execute("SELECT * FROM tests WHERE id=? OR test_code=?", (test_id, test_id)).fetchone()
    conn.close()
    if not t:
        return jsonify({'error': 'Testi nuk u gjet'}), 404
    d = dict(t)
    d['questions'] = json.loads(d['questions'])
    d['grade_levels'] = json.loads(d['grade_levels'])
    return jsonify(d)

@app.route('/api/tests/<test_id>', methods=['DELETE'])
def delete_test(test_id):
    conn = get_db()
    conn.execute("UPDATE tests SET active=0 WHERE id=?", (test_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ── STUDENT SESSION ───────────────────────────────────────────────────────────

@app.route('/api/session/start', methods=['POST'])
def start_session():
    data = request.json
    conn = get_db()
    test = conn.execute("SELECT * FROM tests WHERE (id=? OR test_code=?) AND active=1",
                        (data['test_code'], data['test_code'])).fetchone()
    if not test:
        conn.close()
        return jsonify({'error': 'Kodi i testit është i gabuar ose testi nuk është aktiv'}), 404
    
    sid = str(uuid.uuid4())
    now = time.time()
    questions = json.loads(test['questions'])
    max_score = len(questions) * (10 if len(questions) == 10 else 5)
    
    conn.execute('''INSERT INTO student_sessions 
                    (id, test_id, student_name, student_surname, grade, school, started_at, status, max_score)
                    VALUES (?,?,?,?,?,?,?,'in_progress',?)''',
                 (sid, test['id'], data['name'], data['surname'],
                  data['grade'], data['school'], now, max_score))
    conn.commit()
    
    # Return questions without correct answer
    safe_questions = []
    for i, q in enumerate(questions):
        safe_questions.append({
            'index': i,
            'question': q['question'],
            'options': q['options']
        })
    
    result = {
        'session_id': sid,
        'test_id': test['id'],
        'subject': test['subject'],
        'duration_seconds': test['duration_seconds'],
        'questions': safe_questions,
        'started_at': now
    }
    conn.close()
    return jsonify(result)

@app.route('/api/session/save-answer', methods=['POST'])
def save_answer():
    """Auto-save individual answer"""
    data = request.json
    conn = get_db()
    session = conn.execute("SELECT * FROM student_sessions WHERE id=? AND status='in_progress'",
                           (data['session_id'],)).fetchone()
    if not session:
        conn.close()
        return jsonify({'error': 'Sesioni nuk u gjet'}), 404
    
    answers = json.loads(session['answers'])
    answers[str(data['question_index'])] = data['answer']
    conn.execute("UPDATE student_sessions SET answers=? WHERE id=?",
                 (json.dumps(answers), data['session_id']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/session/submit', methods=['POST'])
def submit_session():
    data = request.json
    conn = get_db()
    session = conn.execute("SELECT * FROM student_sessions WHERE id=?",
                           (data['session_id'],)).fetchone()
    if not session:
        conn.close()
        return jsonify({'error': 'Sesioni nuk u gjet'}), 404
    if session['status'] == 'completed':
        # Return existing result
        result = dict(session)
        result['wrong_questions'] = json.loads(result['wrong_questions'])
        result['correct_questions'] = json.loads(result['correct_questions'])
        result['answers'] = json.loads(result['answers'])
        conn.close()
        return jsonify(result)
    
    test = conn.execute("SELECT * FROM tests WHERE id=?", (session['test_id'],)).fetchone()
    questions = json.loads(test['questions'])
    
    # Use provided answers or saved ones
    answers = data.get('answers', json.loads(session['answers']))
    now = time.time()
    duration = now - session['started_at']
    
    score = 0
    points_per_q = 10 if len(questions) == 10 else 5
    wrong = []
    correct = []
    
    for i, q in enumerate(questions):
        ans = answers.get(str(i))
        if ans == q['correct']:
            score += points_per_q
            correct.append(i)
        else:
            wrong.append({'index': i, 'question': q['question'],
                         'your_answer': ans, 'correct_answer': q['correct'],
                         'options': q['options']})
    
    conn.execute('''UPDATE student_sessions SET 
                    status='completed', submitted_at=?, answers=?, score=?,
                    duration_taken=?, wrong_questions=?, correct_questions=?
                    WHERE id=?''',
                 (now, json.dumps(answers), score, duration,
                  json.dumps(wrong), json.dumps(correct), data['session_id']))
    conn.commit()
    
    result = {
        'session_id': data['session_id'],
        'student_name': session['student_name'],
        'student_surname': session['student_surname'],
        'grade': session['grade'],
        'school': session['school'],
        'subject': test['subject'],
        'score': score,
        'max_score': session['max_score'],
        'duration_taken': duration,
        'wrong_questions': wrong,
        'correct_questions': correct,
        'total_questions': len(questions),
        'submitted_at': now
    }
    conn.close()
    return jsonify(result)

# ── RESULTS / LEADERBOARD ─────────────────────────────────────────────────────

@app.route('/api/results/<test_id>', methods=['GET'])
def get_results(test_id):
    conn = get_db()
    sessions = conn.execute('''
        SELECT ss.*, t.subject FROM student_sessions ss
        JOIN tests t ON ss.test_id = t.id
        WHERE ss.test_id=? AND ss.status='completed'
        ORDER BY ss.score DESC, ss.duration_taken ASC
    ''', (test_id,)).fetchall()
    conn.close()
    result = []
    for i, s in enumerate(sessions):
        d = dict(s)
        d['rank'] = i + 1
        d['wrong_questions'] = json.loads(d['wrong_questions'])
        d['correct_questions'] = json.loads(d['correct_questions'])
        d.pop('answers', None)
        result.append(d)
    return jsonify(result)

@app.route('/api/results/all', methods=['GET'])
def get_all_results():
    conn = get_db()
    sessions = conn.execute('''
        SELECT ss.id, ss.student_name, ss.student_surname, ss.grade, ss.school,
               ss.score, ss.max_score, ss.duration_taken, ss.submitted_at, ss.status,
               ss.wrong_questions, ss.correct_questions,
               t.subject, t.test_code
        FROM student_sessions ss
        JOIN tests t ON ss.test_id = t.id
        WHERE ss.status='completed'
        ORDER BY t.subject, ss.score DESC, ss.duration_taken ASC
    ''').fetchall()
    conn.close()
    result = []
    for s in sessions:
        d = dict(s)
        d['wrong_questions'] = json.loads(d['wrong_questions'])
        d['correct_questions'] = json.loads(d['correct_questions'])
        result.append(d)
    return jsonify(result)

# ── EXPORT ────────────────────────────────────────────────────────────────────

@app.route('/api/export/excel/<test_id>', methods=['GET'])
def export_excel(test_id):
    conn = get_db()
    test = conn.execute("SELECT * FROM tests WHERE id=?", (test_id,)).fetchone()
    sessions = conn.execute('''
        SELECT * FROM student_sessions WHERE test_id=? AND status='completed'
        ORDER BY score DESC, duration_taken ASC
    ''', (test_id,)).fetchall()
    conn.close()
    
    if not test:
        return jsonify({'error': 'Testi nuk u gjet'}), 404
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Rezultatet - {test['subject']}"
    
    # Header styling
    header_fill = PatternFill("solid", fgColor="1a3a5c")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    headers = ['Renditja', 'ID Sesioni', 'Emri', 'Mbiemri', 'Klasa', 'Shkolla',
               'Pikët', 'Pikët Max', '%', 'Kohëzgjatja (min)', 'Pyetje Gabim', 'Pyetje Saktë', 'Data']
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for rank, s in enumerate(sessions, 1):
        duration_min = round(s['duration_taken'] / 60, 1) if s['duration_taken'] else 0
        pct = round(s['score'] / s['max_score'] * 100, 1) if s['max_score'] else 0
        wrong = len(json.loads(s['wrong_questions']))
        correct = len(json.loads(s['correct_questions']))
        submitted = datetime.fromtimestamp(s['submitted_at']).strftime('%d/%m/%Y %H:%M') if s['submitted_at'] else ''
        
        row = [rank, s['id'][:8], s['student_name'], s['student_surname'],
               s['grade'], s['school'], s['score'], s['max_score'], pct,
               duration_min, wrong, correct, submitted]
        
        fill_color = "e8f5e9" if pct >= 50 else "ffebee"
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=rank+1, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal='center')
    
    # Auto width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    fname = f"Rezultatet_{test['subject'].replace(' ','_')}_{test['test_code']}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)

@app.route('/api/export/pdf/<test_id>', methods=['GET'])
def export_pdf(test_id):
    conn = get_db()
    test = conn.execute("SELECT * FROM tests WHERE id=?", (test_id,)).fetchone()
    sessions = conn.execute('''
        SELECT * FROM student_sessions WHERE test_id=? AND status='completed'
        ORDER BY score DESC, duration_taken ASC
    ''', (test_id,)).fetchall()
    conn.close()
    
    if not test:
        return jsonify({'error': 'Testi nuk u gjet'}), 404
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    elements = []
    
    title_style = styles['Title']
    elements.append(Paragraph(f"Gara Komunale e Diturisë 2026", title_style))
    elements.append(Paragraph(f"Lënda: {test['subject']} | Kodi: {test['test_code']}", styles['Normal']))
    elements.append(Spacer(1, 12))
    
    data = [['#', 'ID', 'Emri', 'Mbiemri', 'Klasa', 'Shkolla', 'Pikët', '%', 'Kohë(min)']]
    for rank, s in enumerate(sessions, 1):
        duration_min = round(s['duration_taken'] / 60, 1) if s['duration_taken'] else 0
        pct = round(s['score'] / s['max_score'] * 100, 1) if s['max_score'] else 0
        data.append([rank, s['id'][:8], s['student_name'], s['student_surname'],
                     s['grade'], s['school'], s['score'], f"{pct}%", duration_min])
    
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a3a5c')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f0f4f8')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    output.seek(0)
    
    fname = f"Rezultatet_{test['subject'].replace(' ','_')}_{test['test_code']}.pdf"
    return send_file(output, mimetype='application/pdf', as_attachment=True, download_name=fname)

# ── STATIC ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return send_from_directory('public', 'index.html')

@app.route('/<path:path>')
def static_files(path):
    return send_from_directory('public', path)

if __name__ == '__main__':
    init_db()
    print("✓ Database initialized")
    print("✓ Server starting on http://0.0.0.0:5000")
    app.run(host='0.0.0.0', port=5000, threaded=True, debug=False)
